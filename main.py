import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import webbrowser
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import re


documento = input("digite o nome do documento: ")
# my dict
excel = load_workbook('ALUNOS.xlsx')
listaDeAlunos = excel['Planilha1']

driver = webdriver.Chrome()
driver.get('https://www.intranet.sp.senac.br/home')



def colocar_assim_aparecer(tipo, nome, conteudo, timeout=15):
    try:
        elemento = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((tipo, nome))
        )
        driver.execute_script("document.activeElement.blur();")
        elemento.clear()
        elemento.send_keys(str(conteudo))
        print("✅ Conteúdo inserido com sucesso.")
    except Exception as e:
        pass

def clicar_assim_aparecer(tipo, nome, timeout=15):
    try:
        elemento = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((tipo, nome))
        )
        elemento.click()
        print("✅ Clicou com sucesso.")
    except Exception as e:
        pass

clicar_assim_aparecer(By.XPATH, """/html/body/header/section/nav/ul/li[1]/a""")
clicar_assim_aparecer(By.XPATH, """//*[@id="listagemPessoas"]/div/div[5]/div/div/div[2]/div/div[1]/div[1]/a""")
clicar_assim_aparecer(By.XPATH, """//*[@id="divDetalhesSistema"]/div[2]/div[1]/a""")
driver.switch_to.window(driver.window_handles[-1])

clicar_assim_aparecer(By.XPATH, """//*[@id="app"]/div/div/div[1]/div/div[2]/div[1]/button[2]/div""")
clicar_assim_aparecer(By.XPATH, """//*[@id="item_10282CD6E9BF8D4E23F68971406D4AA4"]/div/li/a/span""")


for row in listaDeAlunos.iter_rows(min_row=2, max_row=listaDeAlunos.max_row, min_col=1, max_col=listaDeAlunos.max_column):
    current_id = row[1].value

    
    colocar_assim_aparecer(By.CSS_SELECTOR, "input[data-v-bc1d237e]", current_id)
    clicar_assim_aparecer(By.XPATH, """//*[@id="app"]/div/div/div[2]/main/div/div/div[2]/div/div[1]/div/button""")
    clicar_assim_aparecer(By.XPATH, """//*[@id="sn-table-desk"]/tbody/tr/td[4]/button""")
    clicar_assim_aparecer(By.XPATH, """//*[@id="app"]/div/div/div[2]/main/div/div/div[2]/div[5]/div[1]/div[2]/button""")
    clicar_assim_aparecer(By.XPATH, """//*[@id="app"]/div/div/div[2]/main/div/div/div[2]/div[5]/div[1]/div[2]/ul/li/a""")
    breakpoint()

    tabela = driver.find_elements(By.XPATH, "//*[@id="app"]/div/div/div[2]/main/div/div/div[2]/div[5]")


    linhas = driver.find_elements(By.XPATH, '//*[@id="sn-table-desk"]/tbody')

    encontrou = False
    for linha in linhas:
        texto = linha.text()
        print(f'{texto}')
        breakpoint()
    
    if documento in texto:
            
