# -*- coding: utf-8 -*-
"""
Verificador de documentos (corrigido e aprimorado)

Recursos:
- Anti-duplicidade (entrada e saída)
- Retentativa automática de IDs que falharam
- Auditoria final (duplicados na entrada e faltantes)
- Colunas fixas F..I com rótulos padronizados e "Pendente" quando ausentes
- "Mostrar Mais → Todos" apenas dentro da página de documentos
- Evita inserir linhas parciais em caso de erro
- RUN_LOG (aba no RESULTADO.xlsx) com o que foi gravado/pulado/erros

Requisitos:
- Selenium (chromedriver compatível com seu Chrome)
- openpyxl
- ALUNOS.xlsx com cabeçalho nas 3 primeiras colunas: Nome | ID | CPF
- RESULTADO.xlsx será criado/atualizado automaticamente
"""

import os
import re
import time
from collections import OrderedDict

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook, Workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException, ElementClickInterceptedException

# ================== FUNÇÕES UTIL ==================
def norm_id(x) -> str:
    """Normaliza ID para dígitos apenas (string)."""
    if x is None:
        return ""
    s = str(x).strip()
    return re.sub(r'\D', '', s)

# ================== ENTRADA ==================
documento = input("digite o nome do documento: ").strip()

# Planilha de alunos (entrada)
excel = load_workbook('ALUNOS.xlsx')
listaDeAlunos = excel['Planilha1']

# Coletar IDs e detectar duplicados na entrada (normalizados)
ids_entrada_norm = []
duplicados_entrada_norm = set()
for r in listaDeAlunos.iter_rows(min_row=2, values_only=True):
    _id = norm_id(r[1] if len(r) > 1 else None)
    if not _id:
        continue
    if _id in ids_entrada_norm:
        duplicados_entrada_norm.add(_id)
    ids_entrada_norm.append(_id)
ids_entrada_unicos_norm = list(OrderedDict.fromkeys(ids_entrada_norm).keys())  # preserva ordem

# ================== PLANILHA DE SAÍDA ==================
OUT_PATH = 'RESULTADO.xlsx'
COL_FIXA_F = "Histórico Escolar do Ensino Fundamental"
COL_FIXA_G = "Declaração de Conclusão do Ensino Fundamental"
COL_FIXA_H = "Declaração de Vacinação Atualizada"
COL_FIXA_I = "Histórico Escolar Parcial para Transferência Externa - Ensino Médio Técnico"

try:
    wb_out = load_workbook(OUT_PATH)
    ws_out = wb_out.active
except Exception:
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.append([
        'Nome', 'ID', 'CPF', 'Total de Doc', 'Status Doc da Pesquisa',
        COL_FIXA_F, COL_FIXA_G, COL_FIXA_H, COL_FIXA_I, 'Documento 1'
    ])

# RUN_LOG para rastrear a execução
if 'RUN_LOG' not in wb_out.sheetnames:
    wb_out.create_sheet('RUN_LOG')
ws_log = wb_out['RUN_LOG']
if ws_log.max_row in (0, 1):
    ws_log.append(['Ordem', 'ID requisitado', 'ID capturado', 'Nome', 'Ação', 'Obs'])

def salvar_resultado_com_retry(max_tries=5, wait=0.4):
    for _ in range(max_tries):
        try:
            wb_out.save(OUT_PATH)
            return True
        except Exception:
            time.sleep(wait)
    print("⚠️ Não consegui salvar RESULTADO.xlsx após várias tentativas (pode estar aberto no Excel/OneDrive).")
    return False

# Conjunto de IDs já processados no RESULTADO.xlsx (normalizados a dígitos)
processados_saida_norm = set()
for row in ws_out.iter_rows(min_row=2, values_only=True):
    _id = norm_id(row[1])
    if _id:
        processados_saida_norm.add(_id)

# Acumuladores
ids_falha_norm = []        # IDs normalizados que falharam nesta rodada
ids_ok_norm = set()        # IDs normalizados concluídos com sucesso (desta execução)
vistos_execucao_norm = set()  # vistos nesta execução (pula duplicados da entrada)

# ================== SELENIUM ==================
driver = webdriver.Chrome()
driver.set_window_size(1440, 900)  # evita layout mobile/accordion
driver.get('https://www.intranet.sp.senac.br/home')

def colocar_assim_aparecer(tipo, nome, conteudo, timeout=15):
    elemento = WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((tipo, nome))
    )
    # Click e limpeza agressiva (resolve casos em que .clear() falha)
    try:
        elemento.click()
    except Exception:
        driver.execute_script("arguments[0].click();", elemento)
    try:
        elemento.clear()
    except Exception:
        pass
    try:
        # Windows: CTRL+A + DEL
        elemento.send_keys(Keys.CONTROL, 'a')
        elemento.send_keys(Keys.DELETE)
    except Exception:
        pass
    elemento.send_keys(str(conteudo))
    print("✅ Campo de busca limpo e preenchido.")
    return elemento

def clicar_assim_aparecer(tipo, nome, timeout=15):
    try:
        elemento = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((tipo, nome))
        )
        try:
            elemento.click()
        except Exception:
            driver.execute_script("arguments[0].click();", elemento)
        print("✅ Clicou com sucesso:", nome)
        return elemento
    except Exception as e:
        print(f"Erro ao clicar no elemento {nome}: {e}")
        raise

# ================== NAVEGAÇÃO INICIAL ==================
clicar_assim_aparecer(By.XPATH, '/html/body/header/section/nav/ul/li[1]/a')
clicar_assim_aparecer(By.XPATH, '//*[@id="listagemPessoas"]/div/div[5]/div/div/div[2]/div/div[1]/div[1]/a')

handles_antes = driver.window_handles.copy()
clicar_assim_aparecer(By.XPATH, '//*[@id="divDetalhesSistema"]/div[2]/div[1]/a')
WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(len(handles_antes) + 1))
driver.switch_to.window(driver.window_handles[-1])

clicar_assim_aparecer(By.XPATH, '//*[@id="app"]/div/div/div[1]/div/div[2]/div[1]/button[2]/div')
clicar_assim_aparecer(By.XPATH, '//*[@id="item_10282CD6E9BF8D4E23F68971406D4AA4"]/div/li/a/span')

# ================== HELPERS ==================
def voltar_para_pesquisa(timeout=20):
    """Volta ao formulário de busca e espera o campo ficar clicável e a grade resetar."""
    try:
        clicar_assim_aparecer(By.XPATH, '//*[@id="app"]/div/div/div[1]/div/div[2]/div[1]/button[2]/div', timeout=timeout)
        clicar_assim_aparecer(By.XPATH, '//*[@id="item_10282CD6E9BF8D4E23F68971406D4AA4"]/div/li/a/span', timeout=timeout)
    except Exception:
        pass
    WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[data-v-bc1d237e]'))
    )
    # aguarda tabela anterior “sumir”, evitando clique na grid ainda carregada
    try:
        WebDriverWait(driver, 4).until_not(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#sn-table-desk tbody tr'))
        )
    except Exception:
        pass

def estou_na_pagina_documentos(timeout=2):
    try:
        WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.XPATH, '//h2[normalize-space()="Documentos do Prontuário Educacional"]'))
        )
        return True
    except Exception:
        return False

def clicar_visualizar_primeira_linha(timeout=20, tentativas=3):
    """Clica no botão Visualizar da primeira linha da grade de busca."""
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, '#sn-table-desk tbody tr'))
    )
    seletores = [
        (By.XPATH, '//*[@id="sn-table-desk"]/tbody/tr[1]/td[4]//button'),
        (By.CSS_SELECTOR, '#sn-table-desk tbody tr:nth-child(1) td:nth-child(4) button[data-cy="read"]'),
        (By.XPATH, '//*[@id="sn-table-desk"]/tbody/tr[1]/td[4]//button[.//img[@alt="Visualizar"]]'),
        (By.XPATH, '//*[@id="sn-table-desk"]/tbody/tr[1]/td[4]//button[.//img[contains(@src,"search-sm-3329cf7e.svg")]]'),
        (By.XPATH, '//*[@id="sn-table-desk"]/tbody/tr/td[4]/button'),
    ]
    for _ in range(tentativas):
        for by, sel in seletores:
            try:
                btn = WebDriverWait(driver, 4).until(EC.element_to_be_clickable((by, sel)))
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                try:
                    btn.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", btn)
                if estou_na_pagina_documentos(timeout=6):
                    print("✅ Entrou na página de documentos.")
                    return True
            except TimeoutException:
                continue
            except (ElementClickInterceptedException, StaleElementReferenceException):
                time.sleep(0.5)
                continue
    raise TimeoutException("Não foi possível clicar em 'Visualizar'.")

def abrir_dropdown_todos_documentos(timeout=12):
    """Abre Mostrar Mais → Todos APENAS na página de documentos."""
    if not estou_na_pagina_documentos():
        return
    try:
        btn = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((
            By.XPATH, '//div[contains(@class,"dropdown")][.//span[normalize-space()="Mostrar Mais"]]//button'
        )))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        try:
            btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", btn)
        menu = WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((
            By.XPATH, '//div[contains(@class,"dropdown")][.//span[normalize-space()="Mostrar Mais"]]//ul[contains(@class,"options")]'
        )))
        opc_todos = menu.find_element(By.XPATH, './/a[normalize-space()="Todos"]')
        try:
            ActionChains(driver).move_to_element(opc_todos).click().perform()
        except Exception:
            driver.execute_script("arguments[0].click();", opc_todos)
    except Exception:
        pass

def garantir_pagina_prontuario(timeout=15):
    if estou_na_pagina_documentos(timeout=2):
        return
    try:
        btn = driver.find_element(By.CSS_SELECTOR, '.collapse-profile-button')
        driver.execute_script("arguments[0].click();", btn)
    except Exception:
        pass
    WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((By.XPATH, '//main//div[contains(@class,"sn-spacing")]'))
    )

def _get_text_by_label(driver, label, timeout=10):
    xp_desk = f'//div[contains(@class,"sn-profile-details")]//h3[normalize-space()="{label}"]/following-sibling::h4[1]'
    try:
        el = WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.XPATH, xp_desk)))
        return el.text.strip()
    except Exception:
        pass
    xp_mob = f'//div[contains(@class,"mobile-profile-details")]//h3[contains(@class,"mobile-title-label")][normalize-space()="{label}"]/following-sibling::h4[contains(@class,"mobile-value-label")][1]'
    el = WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.XPATH, xp_mob)))
    return el.text.strip()

def _doc_base_name(s: str) -> str:
    """Remove ' | Revisão: X' e sufixos ' - ...' para agrupar variantes."""
    if not s:
        return ''
    s = re.sub(r'\s*\|\s*Revis[aã]o:\s*\d+\s*$', '', s, flags=re.IGNORECASE).strip()
    s = s.split(' - ')[0].strip()
    return s

def _parse_docs_desktop(driver):
    docs = []
    rows = driver.find_elements(By.CSS_SELECTOR, '#sn-table-desk tbody tr')
    for r in rows:
        try:
            tds = r.find_elements(By.CSS_SELECTOR, 'td')
            tipo   = tds[0].find_element(By.CSS_SELECTOR, 'span').text.strip()
            origem = tds[1].find_element(By.CSS_SELECTOR, 'span').text.strip()
            data   = tds[2].find_element(By.CSS_SELECTOR, 'span').text.strip()
            docs.append({"tipo": tipo, "origem": origem, "data": data})
        except Exception:
            continue
    return docs

def _parse_docs_mobile(driver):
    docs = []
    titles = driver.find_elements(By.CSS_SELECTOR, '#sn-table-mobile .accordion-button')
    for t in titles:
        tipo = t.text.strip()
        try:
            ancestor = t.find_element(By.XPATH, './ancestor::tr[1]/following-sibling::div[1]')
            origem_el = ancestor.find_element(By.XPATH, './/td[contains(@class,"tbl-accordion-body")][@data-title="Origem :"]/span')
            data_el   = ancestor.find_element(By.XPATH, './/td[contains(@class,"tbl-accordion-body")][@data-title="Data de Análise :"]/span')
            docs.append({"tipo": tipo, "origem": origem_el.text.strip(), "data": data_el.text.strip()})
        except Exception:
            continue
    return docs

def scrape_prontuario(driver, timeout=15):
    garantir_pagina_prontuario(timeout=timeout)
    WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((By.XPATH, '//h2[normalize-space()="Documentos do Prontuário Educacional"]'))
    )
    nome   = _get_text_by_label(driver, 'Nome', timeout)
    cpf    = _get_text_by_label(driver, 'CPF', timeout)
    emplid = _get_text_by_label(driver, 'EMPLID', timeout)
    total  = _get_text_by_label(driver, 'Total de Documentos', timeout)
    documentos = []
    try:
        WebDriverWait(driver, 4).until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#sn-table-desk')))
        documentos = _parse_docs_desktop(driver)
    except Exception:
        try:
            WebDriverWait(driver, 4).until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#sn-table-mobile')))
            documentos = _parse_docs_mobile(driver)
        except Exception:
            documentos = []
    return {"nome": nome, "id": emplid, "cpf": cpf, "total_documentos": total, "documentos": documentos}

# ================== AGRUPAMENTO DE DOCUMENTOS ==================
def so_digitos(s):
    m = re.search(r'\d+', s or '')
    return int(m.group()) if m else 0

def agrega_docs(docs):
    ordem = OrderedDict()
    for d in docs:
        nome_base = _doc_base_name(d.get('tipo', ''))
        if not nome_base:
            continue
        ordem[nome_base] = ordem.get(nome_base, 0) + 1
    saida = []
    for nome, qtd in ordem.items():
        saida.append(f"{nome} ({qtd})" if qtd > 1 else nome)
    return saida, ordem

def status_pesquisa(doc_pesquisado, lista_docs_normalizados):
    alvo = _doc_base_name(doc_pesquisado).casefold()
    return "OK" if any(_doc_base_name(x).casefold() == alvo for x in lista_docs_normalizados) else "Não Encontrado"

def garantir_cabecalho_docs(n_docs_extra):
    headers = [cell.value for cell in ws_out[1]]
    base_fixas = 9
    num_atual = max(0, len(headers) - base_fixas)
    if n_docs_extra > num_atual:
        for i in range(num_atual + 1, n_docs_extra + 1):
            ws_out.cell(row=1, column=base_fixas + i, value=f"Documento {i}")

# ================== LOOP PRINCIPAL (Rodada 1) ==================
for row in listaDeAlunos.iter_rows(min_row=2, max_row=listaDeAlunos.max_row, min_col=1, max_col=listaDeAlunos.max_column):
    current_nome = row[0].value
    current_id_raw = row[1].value
    current_cpf = row[2].value if len(row) >= 3 else None

    nid = norm_id(current_id_raw)
    if not nid:
        print("Linha com ID vazio. Pulando.")
        continue

    # Pular duplicados da entrada já vistos nesta execução
    if nid in vistos_execucao_norm:
        print(f"↪️ Duplicado na entrada (mesma execução): {nid}. Pulando antes do Selenium.")
        continue

    # Pular se já está feito no RESULTADO.xlsx
    if nid in processados_saida_norm:
        print(f"↪️ ID {nid} já consta no RESULTADO.xlsx. Pulando para evitar duplicidade.")
        vistos_execucao_norm.add(nid)
        continue

    try:
        # BUSCA PELO EMPLID
        colocar_assim_aparecer(By.CSS_SELECTOR, 'input[data-v-bc1d237e]', nid)
        clicar_assim_aparecer(By.XPATH, '//*[@id="app"]/div/div/div[2]/main/div/div/div[2]/div/div[1]/div/button')

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#sn-table-desk tbody tr'))
        )

        clicar_visualizar_primeira_linha(timeout=20)
        abrir_dropdown_todos_documentos(timeout=12)

        info = scrape_prontuario(driver, timeout=15)

        nome = info['nome']
        em = info['id']
        em_norm = norm_id(em)
        cpf = info['cpf']
        total_num = so_digitos(info['total_documentos'])
        docs_raw = info['documentos']

        docs_agregados, mapa_contagens = agrega_docs(docs_raw)
        docs_normalizados_lista = list(mapa_contagens.keys())
        status = status_pesquisa(documento, docs_normalizados_lista)

        # Colunas fixas F..I com “Pendente” se ausentes (por nome-base)
        def texto_doc_fixo(rotulo):
            qtd = mapa_contagens.get(_doc_base_name(rotulo), 0)
            if qtd == 0:
                return "Pendente"
            return f"{rotulo} ({qtd})" if qtd > 1 else rotulo

        col_F = texto_doc_fixo(COL_FIXA_F)
        col_G = texto_doc_fixo(COL_FIXA_G)
        col_H = texto_doc_fixo(COL_FIXA_H)
        col_I = texto_doc_fixo(COL_FIXA_I)

        # Extras = tudo que não é uma das 4 colunas fixas (por base name)
        extras = [d for d in docs_agregados
                  if _doc_base_name(d).split(" (")[0] not in {
                      _doc_base_name(COL_FIXA_F),
                      _doc_base_name(COL_FIXA_G),
                      _doc_base_name(COL_FIXA_H),
                      _doc_base_name(COL_FIXA_I),
                  }]

        garantir_cabecalho_docs(len(extras))

        # Checagem final anti-duplicidade (outra execução paralela)
        if em_norm in processados_saida_norm:
            print(f"⚠️ Durante o processamento, o ID {em_norm} apareceu no RESULTADO.xlsx. Ignorando para não duplicar.")
            ws_log.append([ws_log.max_row, nid, em_norm, nome, 'SKIPPED_DUP_RESULT', 'já estava no arquivo'])
            salvar_resultado_com_retry()
        else:
            linha = [nome, em, cpf, total_num, status, col_F, col_G, col_H, col_I] + extras
            ws_out.append(linha)
            ok = salvar_resultado_com_retry()
            if ok:
                processados_saida_norm.add(em_norm)
                ids_ok_norm.add(em_norm)
                vistos_execucao_norm.add(nid)
                ws_log.append([ws_log.max_row, nid, em_norm, nome, 'APPENDED', 'gravado com sucesso'])
                salvar_resultado_com_retry()
            else:
                # não conseguiu salvar → marca para retry
                ids_falha_norm.append(nid)
                ws_log.append([ws_log.max_row, nid, em_norm, nome, 'ERROR_SAVE', 'falha ao salvar; vai para retry'])
                salvar_resultado_com_retry()

        print("====================================")
        print(f"Nome: {nome}")
        print(f"ID: {em}")
        print(f"CPF: {cpf}")
        print(f"Total de Documentos (n°): {total_num}")
        print(f"Status Doc da Pesquisa: {status}")
        print(f"F: {col_F} | G: {col_G} | H: {col_H} | I: {col_I}")
        if extras:
            print("Extras:", ", ".join(extras))

    except Exception as e:
        print(f"Erro no processamento desse ID {nid}: {e}")
        ids_falha_norm.append(nid)
        ws_log.append([ws_log.max_row, nid, None, None, 'ERROR_FETCH', str(e)[:120]])
        salvar_resultado_com_retry()
    finally:
        try:
            voltar_para_pesquisa(timeout=20)
        except Exception:
            try:
                driver.back(); time.sleep(0.8)
                driver.back(); time.sleep(0.8)
                voltar_para_pesquisa(timeout=20)
            except Exception:
                print("Aviso: não consegui voltar à tela de pesquisa automaticamente.")

# ================== LOOP DE RETENTATIVA (Rodada 2) ==================
if ids_falha_norm:
    print("\n🔁 Iniciando rodada de retry para IDs que falharam...\n")
    ids_retry_norm = [i for i in ids_falha_norm if i not in processados_saida_norm]

    for nid in ids_retry_norm:
        try:
            colocar_assim_aparecer(By.CSS_SELECTOR, 'input[data-v-bc1d237e]', nid)
            clicar_assim_aparecer(By.XPATH, '//*[@id="app"]/div/div/div[2]/main/div/div/div[2]/div/div[1]/div/button')

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, '#sn-table-desk tbody tr'))
            )
            clicar_visualizar_primeira_linha(timeout=20)
            abrir_dropdown_todos_documentos(timeout=12)

            info = scrape_prontuario(driver, timeout=15)

            nome = info['nome']
            em = info['id']
            em_norm = norm_id(em)
            cpf = info['cpf']
            total_num = so_digitos(info['total_documentos'])
            docs_raw = info['documentos']

            docs_agregados, mapa_contagens = agrega_docs(docs_raw)
            docs_normalizados_lista = list(mapa_contagens.keys())
            status = status_pesquisa(documento, docs_normalizados_lista)

            def texto_doc_fixo(rotulo):
                qtd = mapa_contagens.get(_doc_base_name(rotulo), 0)
                if qtd == 0:
                    return "Pendente"
                return f"{rotulo} ({qtd})" if qtd > 1 else rotulo

            col_F = texto_doc_fixo(COL_FIXA_F)
            col_G = texto_doc_fixo(COL_FIXA_G)
            col_H = texto_doc_fixo(COL_FIXA_H)
            col_I = texto_doc_fixo(COL_FIXA_I)

            extras = [d for d in docs_agregados
                      if _doc_base_name(d).split(" (")[0] not in {
                          _doc_base_name(COL_FIXA_F),
                          _doc_base_name(COL_FIXA_G),
                          _doc_base_name(COL_FIXA_H),
                          _doc_base_name(COL_FIXA_I),
                      }]
            garantir_cabecalho_docs(len(extras))

            if em_norm in processados_saida_norm:
                print(f"⚠️ ID {em_norm} foi preenchido por outra execução. Pulando.")
                ws_log.append([ws_log.max_row, nid, em_norm, nome, 'RETRY_SKIP_DUP', 'já estava no arquivo'])
                salvar_resultado_com_retry()
            else:
                ws_out.append([nome, em, cpf, total_num, status, col_F, col_G, col_H, col_I] + extras)
                ok = salvar_resultado_com_retry()
                if ok:
                    processados_saida_norm.add(em_norm)
                    ids_ok_norm.add(em_norm)
                    vistos_execucao_norm.add(nid)
                    ws_log.append([ws_log.max_row, nid, em_norm, nome, 'RETRY_OK', 'gravado no retry'])
                    salvar_resultado_com_retry()
                else:
                    ws_log.append([ws_log.max_row, nid, em_norm, nome, 'RETRY_ERROR_SAVE', 'falha ao salvar no retry'])
                    salvar_resultado_com_retry()

        except Exception as e:
            print(f"❌ Retry falhou para {nid}: {e}")
            ws_log.append([ws_log.max_row, nid, None, None, 'RETRY_FAIL', str(e)[:120]])
            salvar_resultado_com_retry()
        finally:
            try:
                voltar_para_pesquisa(timeout=20)
            except Exception:
                try:
                    driver.back(); time.sleep(0.8)
                    driver.back(); time.sleep(0.8)
                    voltar_para_pesquisa(timeout=20)
                except Exception:
                    pass

# ================== AUDITORIA FINAL ==================
esperados = set(ids_entrada_unicos_norm)
existentes = set(processados_saida_norm)

faltantes = sorted(list(esperados - existentes))      # “pulou”/não conseguiu fazer
entrada_dupes = sorted(list(duplicados_entrada_norm)) # duplicados na ALUNOS.xlsx

print("\n===== VERIFICAÇÃO FINAL =====")
print(f"Total na entrada (linhas com ID): {len(ids_entrada_norm)}")
print(f"IDs únicos na entrada: {len(esperados)}")
print(f"Já estavam feitos antes de rodar: {len(processados_saida_norm - ids_ok_norm)}")
print(f"Concluídos nesta execução: {len(ids_ok_norm)}")
print(f"Duplicados na entrada (ALUNOS.xlsx): {len(entrada_dupes)}")
if entrada_dupes:
    print("• IDs duplicados na entrada:", ", ".join(map(str, entrada_dupes)))

print(f"Ficaram faltando (não gravados no RESULTADO.xlsx): {len(faltantes)}")
if faltantes:
    print("• IDs faltantes:", ", ".join(map(str, faltantes)))
    # Gravar guia de auditoria
    try:
        if 'AUDITORIA' not in wb_out.sheetnames:
            wb_out.create_sheet('AUDITORIA')
        ws_aud = wb_out['AUDITORIA']
        ws_aud.delete_rows(1, ws_aud.max_row)  # limpa
        ws_aud.append(['Tipo', 'ID'])
        for _id in entrada_dupes:
            ws_aud.append(['Duplicado na entrada', _id])
        for _id in faltantes:
            ws_aud.append(['Faltante (não processado)', _id])
        salvar_resultado_com_retry()
        print("📄 Guia 'AUDITORIA' atualizada no RESULTADO.xlsx.")
    except Exception as e:
        print(f"Não consegui atualizar guia AUDITORIA: {e}")

# Encerrar o driver de forma limpa
try:
    driver.quit()
except Exception:
    pass

print(f"\n✅ Finalizado. Saída em: {OUT_PATH}")
